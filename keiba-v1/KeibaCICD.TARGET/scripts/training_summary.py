#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
調教データ集計スクリプト

2種類の調教CSV（坂路・コース）を読み込み、馬ごとの調教分類と調教詳細を
タブ区切りファイルで出力する。出力ファイルはTARGET（競馬ソフト）に取り込む。

Usage:
    python training_summary.py --sakamichi sakamichi.csv --course course.csv --date 20251228 --output output.txt
"""

import argparse
import csv
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any


# =============================================================================
# 定数定義
# =============================================================================

# ラップ分類の優先度（点数）
# 記号: + = 加速, - = 減速, = = 同タイム
# 分類: SS > S > A > B > C > D の6段階
# SS = 好タイム + S分類 + 加速or同タイム
CLASSIFICATION_PRIORITY = {
    "SS": 16,
    "S+": 15,
    "S=": 14,
    "S-": 13,
    "A+": 12,
    "A=": 11,
    "A-": 10,
    "B+": 9,
    "B=": 8,
    "B-": 7,
    "C+": 6,
    "C=": 5,
    "C-": 4,
    "D+": 3,
    "D=": 2,
    "D-": 1,
}

# 水木以外でも出力する分類（SS, S, A）
SPECIAL_WEEKDAY_CLASSES = {"SS", "S+", "S-", "S=", "A+", "A-", "A="}

# 好タイム基準（4F、秒）
# 場所: 美浦/栗東、種別: 坂路/コース
GOOD_TIME_THRESHOLD = {
    ("美浦", "坂路"): 52.9,
    ("栗東", "坂路"): 53.9,
    ("美浦", "コース"): 52.2,
    ("栗東", "コース"): 52.2,
}

# 調教タイム分類記号（全角1文字）
TIME_CLASS_SAKAMICHI = "坂"  # 坂路で好タイム
TIME_CLASS_COURSE = "コ"     # コースで好タイム
TIME_CLASS_BOTH = "両"       # 両方で好タイム
TIME_CLASS_NONE = ""         # 好タイムなし（ブランク）


# =============================================================================
# ラップ分類関数
# =============================================================================

def classify_lap(lap2: Optional[float], lap1: Optional[float]) -> Optional[str]:
    """
    Lap2とLap1からラップ分類を判定する
    
    分類ルール（S > A > B > C > D）:
    - S: Lap2が11秒台以下 AND Lap1が11秒台以下（最高）
    - A: Lap1が11秒台以下 AND Lap2が12秒台以上（加速して終い11秒台以下）
    - B: Lap2が12秒台 AND Lap1が12秒台（2F連続12秒台）
    - C: Lap1が12秒台 AND (Lap2が11秒台以下 or 13秒台以上)（終い12秒台）
    - D: Lap1が13秒台以上（軽め）
    
    Args:
        lap2: ラスト2Fラップ（秒）
        lap1: ラスト1Fラップ（秒）
    
    Returns:
        分類文字列（例: "S+", "A-", "B="）またはNone
        + = 加速, - = 減速, = = 同タイム
    """
    if lap2 is None or lap1 is None:
        return None
    
    # 加速/減速/同タイムの判定
    if lap2 > lap1:
        direction = "+"  # 加速
    elif lap2 < lap1:
        direction = "-"  # 減速
    else:
        direction = "="  # 同タイム
    
    # ラップレベルの判定
    # 11秒台以下 = 1, 12秒台 = 2, 13秒台以上 = 3
    def get_level(lap: float) -> int:
        if lap < 12.0:
            return 1  # 11秒台以下（10秒台も含む）
        elif lap < 13.0:
            return 2  # 12秒台
        else:
            return 3  # 13秒台以上
    
    lap2_level = get_level(lap2)
    lap1_level = get_level(lap1)
    
    # 分類判定（優先度順にチェック）
    # S: Lap2が11秒台以下 AND Lap1が11秒台以下
    if lap2_level == 1 and lap1_level == 1:
        return f"S{direction}"
    # A: Lap1が11秒台以下 AND Lap2が12秒台以上
    elif lap1_level == 1 and lap2_level >= 2:
        return f"A{direction}"
    # B: Lap2が12秒台 AND Lap1が12秒台
    elif lap2_level == 2 and lap1_level == 2:
        return f"B{direction}"
    # C: Lap1が12秒台 AND (Lap2が11秒台以下 or 13秒台以上)
    elif lap1_level == 2 and (lap2_level == 1 or lap2_level == 3):
        return f"C{direction}"
    # D: Lap1が13秒台以上
    elif lap1_level == 3:
        return f"D{direction}"
    else:
        # 理論上ここには到達しない
        return None


def get_classification_priority(classification: Optional[str]) -> int:
    """分類の優先度（点数）を取得"""
    if classification is None:
        return 0
    return CLASSIFICATION_PRIORITY.get(classification, 0)


def is_good_time(location: str, source: str, time4f: Optional[float]) -> bool:
    """
    好タイムかどうかを判定する
    
    Args:
        location: 場所（美浦/栗東）
        source: 種別（坂路/コース）
        time4f: 4Fタイム（秒）
    
    Returns:
        好タイムならTrue
    """
    if time4f is None:
        return False
    
    threshold = GOOD_TIME_THRESHOLD.get((location, source))
    if threshold is None:
        return False
    
    return time4f <= threshold


def upgrade_to_ss(classification: Optional[str], has_good_time: bool) -> str:
    """
    好タイムかつS分類（加速or同タイム）の場合、SSに昇格する
    
    Args:
        classification: 元の分類（S+, S=, S- など）
        has_good_time: 好タイムかどうか
    
    Returns:
        SSまたは元の分類
    """
    if classification is None:
        return None
    
    # 好タイム + S分類 + 加速or同タイム → SS
    if has_good_time and classification in ("S+", "S="):
        return "SS"
    
    return classification


def get_time_classification(has_sakamichi_good: bool, has_course_good: bool) -> str:
    """
    調教タイム分類を取得する
    
    Args:
        has_sakamichi_good: 坂路で好タイムがあるか
        has_course_good: コースで好タイムがあるか
    
    Returns:
        調教タイム分類（全角1文字）
    """
    if has_sakamichi_good and has_course_good:
        return TIME_CLASS_BOTH
    elif has_sakamichi_good:
        return TIME_CLASS_SAKAMICHI
    elif has_course_good:
        return TIME_CLASS_COURSE
    else:
        return TIME_CLASS_NONE


# =============================================================================
# 日付判定関数
# =============================================================================

def parse_date(date_str: str) -> datetime:
    """YYYYMMDD形式の文字列をdatetimeに変換"""
    return datetime.strptime(date_str, "%Y%m%d")


def get_training_periods(base_date: datetime) -> Tuple[List[datetime], List[datetime]]:
    """
    基準日から最終追い切り期間と一週前追い切り期間を計算
    
    Args:
        base_date: 基準日（レース日）
    
    Returns:
        (最終追い切り日リスト, 一週前追い切り日リスト)
    """
    # 基準日から遡って直近の水曜・木曜を探す
    # 水曜=2, 木曜=3
    final_dates = []
    one_week_before_dates = []
    
    # 基準日から最大14日前まで遡って水木を探す
    for i in range(1, 15):
        check_date = base_date - timedelta(days=i)
        weekday = check_date.weekday()
        
        if weekday in (2, 3):  # 水曜 or 木曜
            # 最終追い切りの水木を見つける
            if not final_dates:
                # 最初に見つかった水木とその前後1日の水木を最終追い切りとする
                final_dates.append(check_date)
            elif len(final_dates) == 1 and (final_dates[0] - check_date).days == 1:
                # 連続する水木（水→木 or 木→水）
                final_dates.append(check_date)
            elif len(final_dates) >= 1 and (final_dates[-1] - check_date).days >= 6:
                # 一週前の水木
                one_week_before_dates.append(check_date)
                if len(one_week_before_dates) >= 2:
                    break
    
    return final_dates, one_week_before_dates


def get_weekday_display(date: datetime) -> str:
    """曜日表示を取得（水木以外の場合）"""
    weekday_map = {
        0: "(月)",
        1: "(火)",
        4: "(金)",
        5: "(土)",
        6: "(日)",
    }
    return weekday_map.get(date.weekday(), "")


def is_wednesday_or_thursday(date: datetime) -> bool:
    """水曜または木曜かどうか"""
    return date.weekday() in (2, 3)


# =============================================================================
# CSVパーサー
# =============================================================================

def parse_float(value: str) -> Optional[float]:
    """文字列をfloatに変換（失敗時はNone）"""
    try:
        return float(value.strip())
    except (ValueError, AttributeError):
        return None


def read_sakamichi_csv(filepath: Path) -> List[Dict[str, Any]]:
    """
    坂路CSVを読み込む
    
    カラム: 場所,年月日,曜日,馬名,調教師,Time1,Time2,Time3,Time4,Lap4,Lap3,Lap2,Lap1
    """
    records = []
    
    try:
        with open(filepath, "r", encoding="cp932") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                try:
                    record = {
                        "source": "坂路",
                        "location": row.get("場所", ""),
                        "date_str": row.get("年月日", ""),
                        "weekday": row.get("曜日", ""),
                        "horse_name": row.get("馬名", "").strip(),
                        "trainer": row.get("調教師", "").strip(),
                        "time4f": parse_float(row.get("Time1", "")),  # 4F時計
                        "lap2": parse_float(row.get("Lap2", "")),
                        "lap1": parse_float(row.get("Lap1", "")),
                    }
                    
                    # 日付をパース
                    if record["date_str"]:
                        record["date"] = parse_date(record["date_str"])
                    else:
                        continue
                    
                    # 分類を計算
                    record["classification"] = classify_lap(record["lap2"], record["lap1"])
                    
                    # 好タイム判定
                    record["is_good_time"] = is_good_time(
                        record["location"], record["source"], record["time4f"]
                    )
                    
                    # SS昇格判定
                    record["classification"] = upgrade_to_ss(
                        record["classification"], record["is_good_time"]
                    )
                    record["priority"] = get_classification_priority(record["classification"])
                    
                    if record["horse_name"]:
                        records.append(record)
                        
                except Exception as e:
                    print(f"Warning: 坂路CSV {row_num}行目をスキップ: {e}", file=sys.stderr)
                    
    except FileNotFoundError:
        print(f"Error: ファイルが見つかりません: {filepath}", file=sys.stderr)
        raise
    except Exception as e:
        print(f"Error: 坂路CSVの読み込みに失敗: {e}", file=sys.stderr)
        raise
    
    return records


def read_course_csv(filepath: Path) -> List[Dict[str, Any]]:
    """
    コース調教CSVを読み込む
    
    カラム: 場所,コース,回り,年月日,曜日,馬名,調教師,5F,4F,3F,2F,1F,Lap5,Lap4,Lap3,Lap2,Lap1
    """
    records = []
    
    try:
        with open(filepath, "r", encoding="cp932") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                try:
                    record = {
                        "source": "コース",
                        "location": row.get("場所", ""),
                        "course_type": row.get("コース", ""),
                        "direction": row.get("回り", ""),
                        "date_str": row.get("年月日", ""),
                        "weekday": row.get("曜日", ""),
                        "horse_name": row.get("馬名", "").strip(),
                        "trainer": row.get("調教師", "").strip(),
                        "time5f": parse_float(row.get("5F", "")),
                        "time4f": parse_float(row.get("4F", "")),
                        "time1f": parse_float(row.get("1F", "")),
                        "lap2": parse_float(row.get("Lap2", "")),
                        "lap1": parse_float(row.get("Lap1", "")),
                    }
                    
                    # 日付をパース
                    if record["date_str"]:
                        record["date"] = parse_date(record["date_str"])
                    else:
                        continue
                    
                    # 分類を計算
                    record["classification"] = classify_lap(record["lap2"], record["lap1"])
                    
                    # 好タイム判定
                    record["is_good_time"] = is_good_time(
                        record["location"], record["source"], record["time4f"]
                    )
                    
                    # SS昇格判定
                    record["classification"] = upgrade_to_ss(
                        record["classification"], record["is_good_time"]
                    )
                    record["priority"] = get_classification_priority(record["classification"])
                    
                    if record["horse_name"]:
                        records.append(record)
                        
                except Exception as e:
                    print(f"Warning: コースCSV {row_num}行目をスキップ: {e}", file=sys.stderr)
                    
    except FileNotFoundError:
        print(f"Error: ファイルが見つかりません: {filepath}", file=sys.stderr)
        raise
    except Exception as e:
        print(f"Error: コースCSVの読み込みに失敗: {e}", file=sys.stderr)
        raise
    
    return records


# =============================================================================
# 集計処理
# =============================================================================

def aggregate_horse_data(
    sakamichi_records: List[Dict[str, Any]],
    course_records: List[Dict[str, Any]],
    base_date: datetime
) -> Dict[str, Dict[str, Any]]:
    """
    馬ごとにデータを集計する
    
    Returns:
        {馬名: {trainer, final_classification, detail_str}}
    """
    # 最終追い切り期間と一週前追い切り期間を取得
    final_dates, one_week_dates = get_training_periods(base_date)
    
    final_date_set = set(d.date() for d in final_dates)
    one_week_date_set = set(d.date() for d in one_week_dates)
    
    # 全レコードを結合
    all_records = sakamichi_records + course_records
    
    # 馬ごとにグループ化
    horse_records = defaultdict(list)
    for rec in all_records:
        horse_records[rec["horse_name"]].append(rec)
    
    # 馬ごとに集計
    result = {}
    
    for horse_name, records in horse_records.items():
        trainer = records[0]["trainer"] if records else ""
        
        # 期間別にデータを分類
        final_data = []  # 最終追い切り（水木）
        one_week_data = []  # 一週前追い切り（水木）
        other_data = []  # 水木以外で特別分類（S, A）
        
        # 基準日から2週間以内のデータを対象とする
        two_weeks_ago = base_date - timedelta(days=14)
        
        for rec in records:
            rec_date = rec["date"].date()
            is_wed_thu = is_wednesday_or_thursday(rec["date"])
            
            if rec_date in final_date_set:
                final_data.append(rec)
            elif rec_date in one_week_date_set:
                one_week_data.append(rec)
            elif not is_wed_thu and rec["classification"] in SPECIAL_WEEKDAY_CLASSES:
                # 水木以外で特別分類（S, A）の場合
                # 基準日から2週間以内のデータのみ対象
                if rec["date"].date() >= two_weeks_ago.date():
                    other_data.append(rec)
        
        # 各期間で同一日・同一ソースの重複を解決（優先度高い方を採用）
        def dedupe_by_date_source(data: List[Dict]) -> List[Dict]:
            """同一日・同一ソースで優先度最高のものを残す"""
            best = {}
            for rec in data:
                key = (rec["date"].date(), rec["source"])
                if key not in best or rec["priority"] > best[key]["priority"]:
                    best[key] = rec
            return list(best.values())
        
        final_data = dedupe_by_date_source(final_data)
        one_week_data = dedupe_by_date_source(one_week_data)
        other_data = dedupe_by_date_source(other_data)
        
        # 水木がない馬の場合、直近2回のデータを取得
        fallback_data = []
        if not final_data and not one_week_data:
            # 基準日以前のすべてのレコードを日付降順でソート
            all_before_base = [r for r in records if r["date"].date() <= base_date.date() and r["classification"]]
            all_before_base.sort(key=lambda x: (-x["date"].timestamp(), -x["priority"]))
            
            # 同一日・同一ソースの重複を解決
            all_before_base = dedupe_by_date_source(all_before_base)
            
            # 日付降順で直近2回を取得
            all_before_base.sort(key=lambda x: -x["date"].timestamp())
            fallback_data = all_before_base[:2]
        
        # 調教分類の決定
        # 全期間（水木 + 水木以外）で最高優先度を比較して採用
        final_classification = None
        best_priority = 0
        classification_source = None  # "final", "one_week", "other", "fallback"
        
        # 最終追い切り（水木）の最高優先度
        if final_data:
            best_final = max(final_data, key=lambda x: x["priority"])
            if best_final["priority"] > best_priority:
                best_priority = best_final["priority"]
                final_classification = best_final["classification"]
                classification_source = "final"
        
        # 一週前追い切り（水木）の最高優先度
        if one_week_data:
            best_one_week = max(one_week_data, key=lambda x: x["priority"])
            if best_one_week["priority"] > best_priority:
                best_priority = best_one_week["priority"]
                final_classification = best_one_week["classification"]
                classification_source = "one_week"
        
        # 水木以外（土日など）の最高優先度
        if other_data:
            best_other = max(other_data, key=lambda x: x["priority"])
            if best_other["priority"] > best_priority:
                best_priority = best_other["priority"]
                final_classification = best_other["classification"]
                classification_source = "other"
        
        # 水木がない場合のfallbackデータ
        if fallback_data and best_priority == 0:
            best_fallback = max(fallback_data, key=lambda x: x["priority"])
            if best_fallback["priority"] > best_priority:
                best_priority = best_fallback["priority"]
                final_classification = best_fallback["classification"]
                classification_source = "fallback"
        
        if final_classification is None:
            # 分類なしの馬はスキップ
            continue
        
        # 調教タイム分類の計算（全期間で好タイムがあるかどうか）
        all_data = final_data + one_week_data + other_data + fallback_data
        has_sakamichi_good = any(r["is_good_time"] and r["source"] == "坂路" for r in all_data)
        has_course_good = any(r["is_good_time"] and r["source"] == "コース" for r in all_data)
        time_classification = get_time_classification(has_sakamichi_good, has_course_good)
        
        # 調教詳細の作成（好タイムの場合はタイムを括弧で追加）
        def format_record(rec: Dict) -> str:
            """レコードを詳細文字列にフォーマット"""
            if not rec["classification"]:
                return ""
            base = f"{rec['source']}{rec['classification']}"
            if rec.get("is_good_time") and rec.get("time4f"):
                base += f"({rec['time4f']:.1f})"
            return base
        
        detail_parts = []
        
        # 最終追い切り
        if final_data:
            # 優先度順にソート
            final_data.sort(key=lambda x: -x["priority"])
            final_strs = [format_record(rec) for rec in final_data if rec["classification"]]
            if final_strs:
                detail_parts.append("最終" + " ".join(final_strs))
        elif fallback_data:
            # 水木がない場合、直近2回を出力
            fallback_data.sort(key=lambda x: -x["date"].timestamp())
            for rec in fallback_data:
                if rec["classification"]:
                    weekday_str = get_weekday_display(rec["date"])
                    detail_parts.append(f"{weekday_str}{format_record(rec)}")
        else:
            detail_parts.append("最終なし")
        
        # 一週前追い切り（水木がある場合のみ）
        if one_week_data and final_data:
            one_week_data.sort(key=lambda x: -x["priority"])
            one_week_strs = [format_record(rec) for rec in one_week_data if rec["classification"]]
            if one_week_strs:
                detail_parts.append("一週前" + " ".join(one_week_strs))
        elif one_week_data and not final_data and not fallback_data:
            # 最終なしで一週前のみ
            one_week_data.sort(key=lambda x: -x["priority"])
            one_week_strs = [format_record(rec) for rec in one_week_data if rec["classification"]]
            if one_week_strs:
                detail_parts.append("一週前" + " ".join(one_week_strs))
        
        # 水木以外でA分類（fallbackと重複しない場合のみ）
        if not fallback_data:
            for rec in other_data:
                if rec["classification"]:
                    weekday_str = get_weekday_display(rec["date"])
                    detail_parts.append(f"{weekday_str}{format_record(rec)}")
        
        detail_str = " ".join(detail_parts)
        
        result[horse_name] = {
            "trainer": trainer,
            "classification": final_classification,
            "time_class": time_classification,
            "detail": detail_str,
        }
    
    return result


# =============================================================================
# 出力処理
# =============================================================================

def write_output(
    output_path: Path,
    horse_data: Dict[str, Dict[str, Any]]
) -> int:
    """
    タブ区切りファイルを出力
    
    Returns:
        出力した馬数
    """
    with open(output_path, "w", encoding="cp932", newline="") as f:
        # ヘッダー
        f.write("馬名\t調教師\t調教ラップ\t調教タイム\t調教詳細\n")
        
        # データ行
        for horse_name in sorted(horse_data.keys()):
            data = horse_data[horse_name]
            line = f"{horse_name}\t{data['trainer']}\t{data['classification']}\t{data['time_class']}\t{data['detail']}\n"
            f.write(line)
    
    return len(horse_data)


def copy_to_clipboard(text: str) -> bool:
    """
    テキストをクリップボードにコピー（Windows専用）
    
    Returns:
        成功したらTrue
    """
    try:
        # Windowsのclipコマンドを使用
        process = subprocess.Popen(
            ['clip'],
            stdin=subprocess.PIPE,
            shell=True
        )
        process.communicate(text.encode('cp932'))
        return process.returncode == 0
    except Exception as e:
        print(f"Warning: クリップボードへのコピーに失敗: {e}", file=sys.stderr)
        return False


def output_clip_lap(horse_data: Dict[str, Dict[str, Any]]) -> int:
    """
    馬名と調教ラップをクリップボードに出力
    
    Returns:
        出力した馬数
    """
    lines = []
    for horse_name in sorted(horse_data.keys()):
        data = horse_data[horse_name]
        lines.append(f"{horse_name}\t{data['classification']}")
    
    text = "\n".join(lines)
    if copy_to_clipboard(text):
        print("クリップボードにコピーしました（馬名・調教ラップ）")
    
    return len(horse_data)


def output_clip_time(horse_data: Dict[str, Dict[str, Any]]) -> int:
    """
    馬名と調教タイムをクリップボードに出力
    
    Returns:
        出力した馬数
    """
    lines = []
    for horse_name in sorted(horse_data.keys()):
        data = horse_data[horse_name]
        lines.append(f"{horse_name}\t{data['time_class']}")
    
    text = "\n".join(lines)
    if copy_to_clipboard(text):
        print("クリップボードにコピーしました（馬名・調教タイム）")
    
    return len(horse_data)


def output_clip_detail(horse_data: Dict[str, Dict[str, Any]]) -> int:
    """
    馬名と調教詳細をクリップボードに出力
    
    Returns:
        出力した馬数
    """
    lines = []
    for horse_name in sorted(horse_data.keys()):
        data = horse_data[horse_name]
        lines.append(f"{horse_name}\t{data['detail']}")
    
    text = "\n".join(lines)
    if copy_to_clipboard(text):
        print("クリップボードにコピーしました（馬名・調教詳細）")
    
    return len(horse_data)


def write_excel(
    output_path: Path,
    horse_data: Dict[str, Dict[str, Any]]
) -> int:
    """
    Excelファイルを出力
    
    Returns:
        出力した馬数
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("Error: openpyxlがインストールされていません。", file=sys.stderr)
        print("  pip install openpyxl でインストールしてください。", file=sys.stderr)
        sys.exit(1)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "調教データ"
    
    # ヘッダースタイル
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ヘッダー行
    headers = ["馬名", "調教師", "調教ラップ", "調教タイム", "調教詳細"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # データ行
    for row, horse_name in enumerate(sorted(horse_data.keys()), 2):
        data = horse_data[horse_name]
        ws.cell(row=row, column=1, value=horse_name).border = thin_border
        ws.cell(row=row, column=2, value=data['trainer']).border = thin_border
        ws.cell(row=row, column=3, value=data['classification']).border = thin_border
        ws.cell(row=row, column=4, value=data['time_class']).border = thin_border
        ws.cell(row=row, column=5, value=data['detail']).border = thin_border
    
    # 列幅の調整
    ws.column_dimensions['A'].width = 20  # 馬名
    ws.column_dimensions['B'].width = 15  # 調教師
    ws.column_dimensions['C'].width = 10  # 調教ラップ
    ws.column_dimensions['D'].width = 8   # 調教タイム
    ws.column_dimensions['E'].width = 50  # 調教詳細
    
    # 保存
    wb.save(output_path)
    
    return len(horse_data)


# =============================================================================
# メイン処理
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="調教データ集計スクリプト",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  # テキストファイル出力
  python training_summary.py --sakamichi sakamichi.csv --course course.csv --date 20251228 --output output.txt

  # Excelファイル出力
  python training_summary.py -s sakamichi.csv -c course.csv -d 20251228 -o output.xlsx --excel

  # クリップボードに馬名・調教ラップをコピー
  python training_summary.py -s sakamichi.csv -c course.csv -d 20251228 -o output.txt --clip-lap

  # クリップボードに馬名・調教タイムをコピー
  python training_summary.py -s sakamichi.csv -c course.csv -d 20251228 -o output.txt --clip-time

  # クリップボードに馬名・調教詳細をコピー
  python training_summary.py -s sakamichi.csv -c course.csv -d 20251228 -o output.txt --clip-detail
        """
    )
    
    parser.add_argument(
        "-s", "--sakamichi",
        type=Path,
        help="坂路CSVファイルのパス"
    )
    parser.add_argument(
        "-c", "--course",
        type=Path,
        help="コース調教CSVファイルのパス"
    )
    parser.add_argument(
        "-d", "--date",
        type=str,
        required=True,
        help="基準日（YYYYMMDD形式）"
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        required=True,
        help="出力ファイルのパス"
    )
    parser.add_argument(
        "--clip-lap",
        action="store_true",
        help="馬名と調教ラップをクリップボードにコピー"
    )
    parser.add_argument(
        "--clip-time",
        action="store_true",
        help="馬名と調教タイムをクリップボードにコピー"
    )
    parser.add_argument(
        "--clip-detail",
        action="store_true",
        help="馬名と調教詳細をクリップボードにコピー"
    )
    parser.add_argument(
        "--excel",
        action="store_true",
        help="Excelファイル(.xlsx)で出力"
    )
    
    args = parser.parse_args()
    
    # 入力ファイルのチェック
    if not args.sakamichi and not args.course:
        print("Error: --sakamichi または --course のいずれかを指定してください", file=sys.stderr)
        sys.exit(1)
    
    # 基準日のパース
    try:
        base_date = parse_date(args.date)
    except ValueError:
        print(f"Error: 日付形式が不正です: {args.date}（YYYYMMDD形式で指定してください）", file=sys.stderr)
        sys.exit(1)
    
    # データ読み込み
    sakamichi_records = []
    course_records = []
    
    if args.sakamichi:
        if not args.sakamichi.exists():
            print(f"Error: 坂路CSVが見つかりません: {args.sakamichi}", file=sys.stderr)
            sys.exit(1)
        sakamichi_records = read_sakamichi_csv(args.sakamichi)
        print(f"坂路データ読込: {len(sakamichi_records)}件")
    
    if args.course:
        if not args.course.exists():
            print(f"Error: コースCSVが見つかりません: {args.course}", file=sys.stderr)
            sys.exit(1)
        course_records = read_course_csv(args.course)
        print(f"コースデータ読込: {len(course_records)}件")
    
    # 集計
    horse_data = aggregate_horse_data(sakamichi_records, course_records, base_date)
    
    # 出力
    if args.excel:
        # Excel出力
        output_path = args.output.with_suffix('.xlsx') if args.output.suffix != '.xlsx' else args.output
        output_count = write_excel(output_path, horse_data)
        output_file = output_path
    else:
        # テキスト出力
        output_count = write_output(args.output, horse_data)
        output_file = args.output
    
    # クリップボード出力（オプション）
    if args.clip_lap:
        output_clip_lap(horse_data)
    if args.clip_time:
        output_clip_time(horse_data)
    if args.clip_detail:
        output_clip_detail(horse_data)
    
    print(f"\n処理完了:")
    if args.sakamichi:
        print(f"  坂路データ読込: {len(sakamichi_records)}件")
    if args.course:
        print(f"  コースデータ読込: {len(course_records)}件")
    print(f"  出力馬数: {output_count}頭")
    print(f"  出力ファイル: {output_file}")


if __name__ == "__main__":
    main()
